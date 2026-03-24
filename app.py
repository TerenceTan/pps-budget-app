"""
APAC Marketing Budget Tracker — Flask + Google Sheets
Run:  python app.py
Open: http://localhost:5000
"""

import os, json, uuid, base64, io, csv
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, send_file
import gspread
from google.oauth2.service_account import Credentials

# ── CONFIG ────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-production")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

SERVICE_ACCOUNT_FILE = os.environ.get("GOOGLE_CREDS", "credentials.json")
SHEET_ID             = os.environ.get("SHEET_ID", "13TMeZ3pqdUQr2WRMG5G70xchZKfmiPVOShZChqbUsv4")
ADMIN_MARKET         = "APAC"

MARKETS  = ["CN","HKG","ID","IN","MN","MY","PH","SG","TH","TW","VN","TW/SG/MY/MN"]
QUARTERS = ["Q1","Q2","Q3","Q4"]

# Invoice storage on disk (not in Google Sheets — avoids cell size limits)
INVOICE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoices")
os.makedirs(INVOICE_DIR, exist_ok=True)

def save_invoice_to_disk(data_url, entry_id, filename):
    """Save a base64 data URL to disk, return the stored filename."""
    try:
        header, b64 = data_url.split(",", 1)
        data = base64.b64decode(b64)
        # Sanitise filename
        safe_name = "".join(c for c in filename if c.isalnum() or c in "._- ")[:100]
        stored_name = f"{entry_id}_{uuid.uuid4().hex[:6]}_{safe_name}"
        path = os.path.join(INVOICE_DIR, stored_name)
        with open(path, "wb") as f:
            f.write(data)
        return stored_name
    except Exception:
        return None

def get_invoice_path(stored_name):
    """Get full path for a stored invoice."""
    path = os.path.join(INVOICE_DIR, stored_name)
    if os.path.exists(path):
        return path
    return None

TAB_BUDGETS    = "Budgets"
TAB_CHANNELS   = "Channels"
TAB_ACTIVITIES = "Activities"
TAB_ENTRIES    = "Entries"
TAB_MAPPING    = "ChannelMapping"

ENTRY_HEADERS = [
    "id","country","quarter","month","channel_id","channel_name",
    "activity_id","activity_name",
    "bu","finance_cat","marketing_cat","description",
    "planned","confirmed","actual",
    "jira","vendor","notes",
    "approved","invoice_names","invoice_data",
    "entered_by","created_at","updated_at"
]

MAPPING_HEADERS = ["channel_keyword","bu","finance_cat","marketing_cat","updated_by","updated_at"]

DEFAULT_MAPPING = [
    ("performance",  "Marketing : Programmatic - 613000009XXX",          "PPC",                  "PPC / Search"),
    ("ppc",          "Marketing : Programmatic - 613000009XXX",          "PPC",                  "PPC / Search"),
    ("programmatic", "Marketing : Programmatic - 613000009XXX",          "Programmatic",         "Programmatic"),
    ("affiliate",    "Marketing : Affiliate - 613000002XXX",             "Affiliate",            "Affiliate- CPA & FF"),
    ("paid social",  "Marketing : Paid Social / YouTube - 613000004XXX", "Paid Social-Meta",     "Paid Social"),
    ("social",       "Marketing : Paid Social / YouTube - 613000004XXX", "Paid Social-Meta",     "Paid Social"),
    ("youtube",      "Marketing : Paid Social / YouTube - 613000004XXX", "Paid Social-Youtube",  "YouTube"),
    ("brand",        "Marketing : Local Brand - 613000024XXX",           "Campaigns/Promotions", "Brand / OOH"),
    ("event",        "Marketing : Local Brand - 613000024XXX",           "Event",                "Events & Sponsorship"),
    ("influencer",   "Marketing : Local Brand - 613000024XXX",           "Influencer/KOL",       "Influencer / KOL"),
    ("kol",          "Marketing : Local Brand - 613000024XXX",           "Influencer/KOL",       "Influencer / KOL"),
    ("premium",      "Marketing : Premium - 613000019XXX",               "Premium",              "Premium Partners"),
    ("partner",      "Marketing : Partners - 613000022XXX",              "Partner",              "Premium Partners"),
    ("raf",          "Marketing - Refer a friend - 613000003XXX",        "RAF",                  "Refer a Friend"),
    ("refer",        "Marketing - Refer a friend - 613000003XXX",        "RAF",                  "Refer a Friend"),
    ("mar tech",     "Marketing : Marketing technology",                  "Marketing Technology", "Technology"),
    ("technology",   "Marketing : Marketing technology",                  "Marketing Technology", "Technology"),
    ("seo",          "Marketing : Local Brand - 613000024XXX",           "Local SEO",            "SEO"),
    ("amf1",         "Marketing : Partners - 613000022XXX",              "AMF1 Activation",      "AMF1"),
]

# ── SHEETS HELPERS ────────────────────────────────────────────
# In-memory cache — reduces Sheets API calls significantly
import time
_sheet_cache = {}
CACHE_TTL = 30  # seconds

# Persistent gspread client + spreadsheet (avoids re-auth on every call)
_gc = None
_gc_ts = 0
_GC_TTL = 600  # re-auth every 10 minutes

def get_gc():
    global _gc, _gc_ts
    now = time.time()
    if _gc is None or (now - _gc_ts) > _GC_TTL:
        try:
            creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            _gc = gspread.authorize(creds)
            _gc_ts = now
        except Exception as e:
            _gc = None
            raise e
    return _gc

_sh = None
_sh_ts = 0

def get_spreadsheet():
    global _sh, _sh_ts
    now = time.time()
    if _sh is None or (now - _sh_ts) > _GC_TTL:
        try:
            gc = get_gc()
            _sh = gc.open_by_key(SHEET_ID)
            _sh_ts = now
        except Exception as e:
            _sh = None
            raise e
    return _sh

def get_sheet(tab):
    global _sh, _sh_ts
    sh = get_spreadsheet()
    try:
        return sh.worksheet(tab)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab, rows=2000, cols=26)
        _init_headers(ws, tab)
        return ws
    except Exception:
        # Stale spreadsheet object — force refresh and retry
        _sh = None
        _sh_ts = 0
        sh = get_spreadsheet()
        try:
            return sh.worksheet(tab)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=tab, rows=2000, cols=26)
            _init_headers(ws, tab)
            return ws

def get_records_cached(tab):
    """Get all records with caching. Mutations call invalidate_cache(tab)."""
    now = time.time()
    if tab in _sheet_cache:
        ts, data = _sheet_cache[tab]
        if now - ts < CACHE_TTL:
            return data
    try:
        data = get_sheet(tab).get_all_records()
    except Exception:
        data = []
    _sheet_cache[tab] = (now, data)
    return data

def invalidate_cache(tab):
    _sheet_cache.pop(tab, None)

def rows_for_cached(tab, **filters):
    rows = get_records_cached(tab)
    for k, v in filters.items():
        rows = [r for r in rows if str(r.get(k,"")) == str(v)]
    return rows

def _init_headers(ws, tab):
    hdrs = {
        TAB_BUDGETS:    ["id","country","quarter","total_budget","updated_at"],
        TAB_CHANNELS:   ["id","country","quarter","name","budget","sort_order","created_at"],
        TAB_ACTIVITIES: ["id","channel_id","country","quarter","name","sort_order","created_at"],
        TAB_ENTRIES:    ENTRY_HEADERS,
        TAB_MAPPING:    MAPPING_HEADERS,
    }
    if tab in hdrs:
        ws.append_row(hdrs[tab])
        ws.format("1:1", {
            "backgroundColor": {"red":0.11,"green":0.31,"blue":0.24},
            "textFormat": {"foregroundColor":{"red":1,"green":1,"blue":1},"bold":True}
        })

def rows_for(tab, **filters):
    try:
        rows = get_sheet(tab).get_all_records()
    except Exception:
        rows = []
    for k, v in filters.items():
        rows = [r for r in rows if str(r.get(k,"")) == str(v)]
    return rows

# ── AUTH DECORATORS ───────────────────────────────────────────
def require_login(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            return jsonify({"error":"Not authenticated"}), 401
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user") != ADMIN_MARKET:
            return jsonify({"error":"Admin required"}), 403
        return f(*args, **kwargs)
    return decorated

def check_country_access(country):
    u = session.get("user")
    if u != ADMIN_MARKET and u != country:
        return False
    return True

# ── PAGES ─────────────────────────────────────────────────────
@app.route("/")
def index():
    if not session.get("user"):
        return render_template("login.html", markets=MARKETS)
    return render_template("app.html",
        user=session["user"],
        is_admin=(session["user"] == ADMIN_MARKET),
        markets=MARKETS, quarters=QUARTERS
    )

@app.route("/login", methods=["POST"])
def login():
    market = request.form.get("market","").strip()
    if market in MARKETS + [ADMIN_MARKET]:
        session["user"] = market
    return redirect("/")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ── BUDGET API ────────────────────────────────────────────────
@app.route("/api/budget/<country>/<quarter>")
@require_login
def api_get_budget(country, quarter):
    if not check_country_access(country):
        return jsonify({"error":"Forbidden"}), 403
    try:
        brows = rows_for_cached(TAB_BUDGETS, country=country, quarter=quarter)
        total = float(brows[0]["total_budget"]) if brows else 0

        channels = sorted([
            {"id":r["id"],"name":r["name"],"budget":float(r["budget"] or 0),"sort_order":int(r.get("sort_order") or 0)}
            for r in rows_for_cached(TAB_CHANNELS, country=country, quarter=quarter)
        ], key=lambda x: x["sort_order"])

        # Attach activities to each channel
        all_acts = rows_for_cached(TAB_ACTIVITIES, country=country, quarter=quarter)
        for ch in channels:
            ch["activities"] = sorted([
                {"id":a["id"],"name":a["name"],"sort_order":int(a.get("sort_order") or 0)}
                for a in all_acts if a["channel_id"] == ch["id"]
            ], key=lambda x: x["sort_order"])

        # Load mapping for auto-fill
        try:
            mrows = get_records_cached(TAB_MAPPING)
            if not mrows:
                _seed_mapping()
                invalidate_cache(TAB_MAPPING)
                mrows = get_records_cached(TAB_MAPPING)
            mapping = [{"channel_keyword":r["channel_keyword"],"bu":r["bu"],"finance_cat":r["finance_cat"],"marketing_cat":r["marketing_cat"]} for r in mrows if r.get("channel_keyword")]
        except Exception:
            mapping = [{"channel_keyword":kw,"bu":bu,"finance_cat":fc,"marketing_cat":mc} for kw,bu,fc,mc in DEFAULT_MAPPING]

        return jsonify({"total":total, "channels":channels, "mapping":mapping})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Budget load failed: {str(e)}"}), 500

def _seed_mapping():
    ws = get_sheet(TAB_MAPPING)
    now = datetime.utcnow().isoformat()
    for kw,bu,fc,mc in DEFAULT_MAPPING:
        ws.append_row([kw,bu,fc,mc,"system",now])

# ── ACTIVITIES API ────────────────────────────────────────────
@app.route("/api/activities", methods=["POST"])
@require_login
@require_admin
def api_add_activity():
    d = request.get_json()
    existing = rows_for(TAB_ACTIVITIES, channel_id=d["channel_id"])
    act_id = "act_" + str(uuid.uuid4())[:8]
    get_sheet(TAB_ACTIVITIES).append_row([
        act_id, d["channel_id"], d["country"], d["quarter"],
        d["name"], len(existing), datetime.utcnow().isoformat()
    ])
    invalidate_cache(TAB_ACTIVITIES)
    return jsonify({"id":act_id,"name":d["name"],"sort_order":len(existing)})

@app.route("/api/activities/<act_id>", methods=["PUT"])
@require_login
@require_admin
def api_update_activity(act_id):
    d = request.get_json()
    ws = get_sheet(TAB_ACTIVITIES)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==act_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    r = rows[idx]
    ws.update(f"A{idx+2}:G{idx+2}", [[act_id, r["channel_id"], r["country"], r["quarter"], d.get("name",r["name"]), r.get("sort_order",0), r.get("created_at","")]])
    invalidate_cache(TAB_ACTIVITIES)
    return jsonify({"ok":True})

@app.route("/api/activities/<act_id>", methods=["DELETE"])
@require_login
@require_admin
def api_delete_activity(act_id):
    ws = get_sheet(TAB_ACTIVITIES)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==act_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    ws.delete_rows(idx+2)
    invalidate_cache(TAB_ACTIVITIES)
    return jsonify({"ok":True})

# ── MAPPING API ──────────────────────────────────────────────
@app.route("/api/mapping")
@require_login
def api_get_mapping():
    try:
        mrows = get_records_cached(TAB_MAPPING)
        if not mrows:
            _seed_mapping()
            invalidate_cache(TAB_MAPPING)
            mrows = get_records_cached(TAB_MAPPING)
        mapping = [{"channel_keyword":r["channel_keyword"],"bu":r["bu"],"finance_cat":r["finance_cat"],"marketing_cat":r["marketing_cat"]} for r in mrows if r.get("channel_keyword")]
    except Exception:
        mapping = [{"channel_keyword":kw,"bu":bu,"finance_cat":fc,"marketing_cat":mc} for kw,bu,fc,mc in DEFAULT_MAPPING]
    return jsonify(mapping)

# ── MAPPING SAVE ──────────────────────────────────────────────
@app.route("/api/mapping", methods=["POST"])
@require_login
@require_admin
def api_save_mapping():
    mappings = request.get_json()
    ws = get_sheet(TAB_MAPPING)
    ws.clear()
    ws.append_row(MAPPING_HEADERS)
    ws.format("1:1", {"backgroundColor":{"red":0.11,"green":0.31,"blue":0.24},"textFormat":{"foregroundColor":{"red":1,"green":1,"blue":1},"bold":True}})
    now = datetime.utcnow().isoformat()
    for m in mappings:
        ws.append_row([m.get("channel_keyword",""), m.get("bu",""), m.get("finance_cat",""), m.get("marketing_cat",""), session["user"], now])
    invalidate_cache(TAB_MAPPING)
    return jsonify({"ok":True, "count":len(mappings)})

@app.route("/api/budget/<country>/<quarter>", methods=["POST"])
@require_login
@require_admin
def api_save_budget(country, quarter):
    total = float(request.get_json().get("total", 0))
    ws  = get_sheet(TAB_BUDGETS)
    rows = ws.get_all_records()
    now = datetime.utcnow().isoformat()
    idx = next((i for i,r in enumerate(rows) if r["country"]==country and r["quarter"]==quarter), None)
    if idx is not None:
        ws.update(f"A{idx+2}:E{idx+2}", [[rows[idx]["id"], country, quarter, total, now]])
    else:
        ws.append_row([str(uuid.uuid4())[:8], country, quarter, total, now])
    invalidate_cache(TAB_BUDGETS)
    return jsonify({"ok":True})

# ── CHANNELS API ──────────────────────────────────────────────
@app.route("/api/channels", methods=["POST"])
@require_login
@require_admin
def api_add_channel():
    d = request.get_json()
    existing = rows_for(TAB_CHANNELS, country=d["country"], quarter=d["quarter"])
    ch_id = "ch_" + str(uuid.uuid4())[:8]
    get_sheet(TAB_CHANNELS).append_row([
        ch_id, d["country"], d["quarter"], d["name"],
        float(d.get("budget",0)), len(existing), datetime.utcnow().isoformat()
    ])
    invalidate_cache(TAB_CHANNELS)
    return jsonify({"id":ch_id,"name":d["name"],"budget":float(d.get("budget",0)),"sort_order":len(existing)})

@app.route("/api/channels/<ch_id>", methods=["PUT"])
@require_login
@require_admin
def api_update_channel(ch_id):
    d = request.get_json()
    ws = get_sheet(TAB_CHANNELS)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==ch_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    r = rows[idx]
    ws.update(f"A{idx+2}:G{idx+2}", [[
        ch_id, r["country"], r["quarter"],
        d.get("name", r["name"]), float(d.get("budget", r["budget"])),
        r.get("sort_order",0), r.get("created_at","")
    ]])
    invalidate_cache(TAB_CHANNELS)
    return jsonify({"ok":True})

@app.route("/api/channels/<ch_id>", methods=["DELETE"])
@require_login
@require_admin
def api_delete_channel(ch_id):
    ws = get_sheet(TAB_CHANNELS)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==ch_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    ws.delete_rows(idx+2)
    invalidate_cache(TAB_CHANNELS)
    return jsonify({"ok":True})

# ── ENTRIES API ───────────────────────────────────────────────
@app.route("/api/entries/<country>/<quarter>")
@require_login
def api_get_entries(country, quarter):
    if not check_country_access(country):
        return jsonify({"error":"Forbidden"}), 403
    try:
        rows = rows_for_cached(TAB_ENTRIES, country=country, quarter=quarter)
        entries = []
        for r in rows:
            entries.append({
                "id": r["id"], "country": r["country"], "quarter": r["quarter"],
                "month": r["month"], "channel_id": r["channel_id"], "channel_name": r["channel_name"],
                "bu": r["bu"], "finance_cat": r["finance_cat"], "marketing_cat": r["marketing_cat"],
                "activity_id": r.get("activity_id",""), "activity_name": r.get("activity_name",""),
                "description": r["description"],
                "planned":   float(r["planned"]   or 0),
                "confirmed": float(r["confirmed"] or 0),
                "actual":    float(r["actual"]    or 0),
                "jira": r["jira"], "vendor": r["vendor"], "notes": r["notes"],
                "approved": str(r["approved"]).lower() == "true",
                "invoice_names": json.loads(r["invoice_names"]) if r.get("invoice_names") else [],
                "entered_by": r["entered_by"], "updated_at": r["updated_at"],
            })
        return jsonify(entries)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Entries load failed: {str(e)}"}), 500

@app.route("/api/entries", methods=["POST"])
@require_login
def api_add_entry():
    d = request.get_json()
    country = d.get("country","")
    if not check_country_access(country):
        return jsonify({"error":"Forbidden"}), 403
    entry_id = "e_" + str(uuid.uuid4())[:10]
    now = datetime.utcnow().isoformat()

    # Save invoice files to disk, store only filenames in Sheets
    inv_names = d.get("invoice_names", [])
    inv_data_urls = d.get("invoice_data", [])
    stored_files = []
    for i, name in enumerate(inv_names):
        if i < len(inv_data_urls) and inv_data_urls[i]:
            stored = save_invoice_to_disk(inv_data_urls[i], entry_id, name)
            if stored:
                stored_files.append(stored)
        # else: name without data (shouldn't happen on add, but handle gracefully)

    get_sheet(TAB_ENTRIES).append_row([
        entry_id, country,
        d.get("quarter",""), d.get("month",""),
        d.get("channel_id",""), d.get("channel_name",""),
        d.get("activity_id",""), d.get("activity_name",""),
        d.get("bu",""), d.get("finance_cat",""), d.get("marketing_cat",""),
        d.get("description",""),
        float(d.get("planned") or 0), float(d.get("confirmed") or 0), float(d.get("actual") or 0),
        d.get("jira",""), d.get("vendor",""), d.get("notes",""),
        str(d.get("approved", False)),
        json.dumps(inv_names),
        json.dumps(stored_files),
        session["user"], now, now
    ])
    invalidate_cache(TAB_ENTRIES)
    return jsonify({"id":entry_id,"ok":True})

@app.route("/api/entries/<entry_id>", methods=["PUT"])
@require_login
def api_update_entry(entry_id):
    d = request.get_json()
    ws = get_sheet(TAB_ENTRIES)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==entry_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    r = rows[idx]
    if not check_country_access(r["country"]):
        return jsonify({"error":"Forbidden"}), 403

    approved = d.get("approved", str(r["approved"]).lower()=="true")
    jira = d.get("jira", r["jira"])
    if approved and not jira:
        return jsonify({"error":"JIRA link required before approving"}), 400

    inv_names_new = d.get("invoice_names", None)
    inv_data_new  = d.get("invoice_data", None)

    # Existing stored file references from Sheets
    existing_names = json.loads(r.get("invoice_names") or "[]")
    existing_files = json.loads(r.get("invoice_data")  or "[]")

    if inv_names_new is not None:
        new_data_urls = inv_data_new if inv_data_new else []

        final_names = []
        final_files = []
        for name in inv_names_new:
            # Check if it's a surviving existing invoice
            if name in existing_names:
                old_idx = existing_names.index(name)
                final_names.append(name)
                if old_idx < len(existing_files):
                    final_files.append(existing_files[old_idx])
                else:
                    final_files.append("")
                existing_names[old_idx] = None  # mark used
            elif new_data_urls:
                # New upload — save to disk
                data_url = new_data_urls.pop(0)
                stored = save_invoice_to_disk(data_url, entry_id, name)
                final_names.append(name)
                final_files.append(stored or "")
            else:
                final_names.append(name)
                final_files.append("")

        inv_names = json.dumps(final_names)
        inv_data  = json.dumps(final_files)
    else:
        inv_names = json.dumps(existing_names)
        inv_data  = json.dumps(existing_files)
    now = datetime.utcnow().isoformat()

    ws.update(f"A{idx+2}:W{idx+2}", [[
        entry_id, r["country"],
        d.get("quarter", r["quarter"]), d.get("month", r["month"]),
        d.get("channel_id", r["channel_id"]), d.get("channel_name", r["channel_name"]),
        d.get("activity_id", r.get("activity_id","")), d.get("activity_name", r.get("activity_name","")),
        d.get("bu", r["bu"]), d.get("finance_cat", r["finance_cat"]),
        d.get("marketing_cat", r["marketing_cat"]),
        d.get("description", r["description"]),
        float(d.get("planned", r["planned"]) or 0),
        float(d.get("confirmed", r["confirmed"]) or 0),
        float(d.get("actual", r["actual"]) or 0),
        jira, d.get("vendor", r["vendor"]), d.get("notes", r["notes"]),
        str(approved), inv_names, inv_data,
        r["entered_by"], r["created_at"], now
    ]])
    invalidate_cache(TAB_ENTRIES)
    return jsonify({"ok":True})

@app.route("/api/entries/<entry_id>", methods=["DELETE"])
@require_login
def api_delete_entry(entry_id):
    user = session["user"]
    ws = get_sheet(TAB_ENTRIES)
    rows = ws.get_all_records()
    idx = next((i for i,r in enumerate(rows) if r["id"]==entry_id), None)
    if idx is None: return jsonify({"error":"Not found"}), 404
    # Admin can delete anything; market users can only delete their own
    if user != ADMIN_MARKET and rows[idx]["country"] != user:
        return jsonify({"error":"Forbidden"}), 403
    ws.delete_rows(idx+2)
    invalidate_cache(TAB_ENTRIES)
    return jsonify({"ok":True})

# ── TEMPLATE DOWNLOADS ────────────────────────────────────────
@app.route("/api/channel_template")
@require_login
def api_channel_template():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Country', 'Quarter', 'Channel Name', 'Budget (USD)'])
    w.writerow(['TH', 'Q1', 'Performance Marketing', 50000])
    w.writerow(['TH', 'Q1', 'Affiliate', 30000])
    w.writerow(['TH', 'Q1', 'Paid Social', 25000])
    w.writerow(['TH', 'Q1', 'Regional Marketing', 20000])
    w.writerow(['SG', 'Q1', 'Performance Marketing', 40000])
    w.writerow(['SG', 'Q1', 'Affiliate', 20000])
    w.writerow(['MY', 'Q2', 'Performance Marketing', 35000])
    w.writerow(['MY', 'Q2', 'Paid Social', 15000])
    return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True,
                     download_name='channel_budget_template.csv')

@app.route("/api/budget_template")
@require_login
def api_budget_template():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Country', 'Quarter', 'Total Budget (USD)'])
    for country in ['TH','SG','MY','CN','HKG','ID','IN','VN','PH','TW','MN']:
        for q in ['Q1','Q2','Q3','Q4']:
            w.writerow([country, q, 0])
    return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True,
                     download_name='country_budget_template.csv')

# ── INVOICE DOWNLOAD ──────────────────────────────────────────
@app.route("/api/invoice/<entry_id>/<int:inv_idx>")
@require_login
def api_invoice(entry_id, inv_idx):
    rows = get_sheet(TAB_ENTRIES).get_all_records()
    r = next((row for row in rows if row["id"]==entry_id), None)
    if not r: return "Not found", 404
    names = json.loads(r.get("invoice_names") or "[]")
    datas = json.loads(r.get("invoice_data")  or "[]")
    if inv_idx >= len(datas): return "Not found", 404
    name = names[inv_idx] if inv_idx < len(names) else f"invoice_{inv_idx}"
    stored = datas[inv_idx]

    # New format: stored filename on disk
    if stored and not stored.startswith("data:"):
        path = get_invoice_path(stored)
        if path:
            import mimetypes
            mime = mimetypes.guess_type(name)[0] or "application/octet-stream"
            return send_file(path, mimetype=mime, as_attachment=True, download_name=name)
        return "File not found on disk", 404

    # Legacy format: base64 data URL stored in Sheets
    if stored and stored.startswith("data:"):
        try:
            header, b64 = stored.split(",", 1)
            mime = header.split(";")[0].replace("data:","")
            return send_file(io.BytesIO(base64.b64decode(b64)), mimetype=mime,
                             as_attachment=True, download_name=name)
        except Exception:
            return "Invalid invoice data", 500

    return "No file data", 404

# ── BULK UPLOAD: EXCEL PARSER ────────────────────────────────
@app.route("/api/parse_bulk", methods=["POST"])
@require_login
def api_parse_bulk():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files['file']
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Please upload an .xlsx or .xls file"}), 400
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True, read_only=True)
        ws = wb.active
        rows_out = []
        all_rows = list(ws.iter_rows(values_only=True))
        # Detect header
        start = 0
        if all_rows:
            first = [str(v or '').lower() for v in all_rows[0]]
            if any(k in ' '.join(first) for k in ['month','vendor','planned','campaign']):
                start = 1
        for row in all_rows[start:]:
            vals = [str(v).strip() if v is not None else '' for v in row]
            if not any(vals): continue
            def safe_float(s):
                try: return float(str(s).replace(',','').replace('$','').strip() or 0)
                except: return 0
            rows_out.append({
                'month':     vals[0] if len(vals)>0 else '',
                'campaign':  vals[1] if len(vals)>1 else '',
                'desc':      vals[2] if len(vals)>2 else '',
                'vendor':    vals[3] if len(vals)>3 else '',
                'planned':   safe_float(vals[4]) if len(vals)>4 else 0,
                'confirmed': safe_float(vals[5]) if len(vals)>5 else 0,
                'actual':    safe_float(vals[6]) if len(vals)>6 else 0,
                'jira':      vals[7] if len(vals)>7 else '',
                'notes':     vals[8] if len(vals)>8 else '',
            })
        wb.close()
        return jsonify({"rows": rows_out})
    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 400

# ── BULK UPLOAD: ENTRY TEMPLATE ──────────────────────────────
@app.route("/api/bulk_template")
@require_login
def api_bulk_template():
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Month','Activity','Description','Vendor',
                'Planned (USD)','Confirmed (USD)','Actual (USD)','JIRA Task','Notes'])
    w.writerow(['Jul 2025','Google Search Q1','Brand keywords','Google',5000,'',3200,'MKT-001',''])
    w.writerow(['Aug 2025','Meta Retargeting','Retargeting campaign','Meta',3000,'',2800,'MKT-002',''])
    w.writerow(['Sep 2025','Bing Display','Display network','Bing',2000,'',1500,'','Paused mid-month'])
    return send_file(
        io.BytesIO(out.getvalue().encode('utf-8-sig')),
        mimetype='text/csv', as_attachment=True,
        download_name='bulk_entry_template.csv'
    )

# ── BULK CHANNEL IMPORT ────────────────────────────────────────
@app.route("/api/import/channels", methods=["POST"])
@require_login
@require_admin
def api_import_channels():
    """
    Accepts CSV or XLSX with columns: Country, Quarter, Channel Name, Budget
    Creates channels (and budget records if missing) for each row.
    Returns the actual rows that were saved for preview.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files['file']
    fname = f.filename.lower()

    # Parse file into rows
    parsed_rows = []
    try:
        if fname.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return jsonify({"error": "Empty file"}), 400
            first = [str(v or '').lower() for v in all_rows[0]]
            start = 1 if any(k in ' '.join(first) for k in ['country','quarter','channel','name']) else 0
            for row in all_rows[start:]:
                vals = [str(v).strip() if v is not None else '' for v in row]
                if any(vals):
                    parsed_rows.append(vals)
        else:
            content = f.read().decode('utf-8-sig')
            lines = [l.strip() for l in content.splitlines() if l.strip()]
            if not lines:
                return jsonify({"error": "Empty file"}), 400
            first = lines[0].lower()
            has_header = 'country' in first or 'quarter' in first or 'channel' in first
            data_lines = lines[1:] if has_header else lines
            import csv as _csv
            parsed_rows = list(_csv.reader(data_lines))
    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 400

    saved, skipped = 0, 0
    saved_rows = []
    ws_ch = get_sheet(TAB_CHANNELS)
    ws_bud = get_sheet(TAB_BUDGETS)
    existing_ch = ws_ch.get_all_records()
    existing_bud = ws_bud.get_all_records()
    now = datetime.utcnow().isoformat()

    for row in parsed_rows:
        if len(row) < 3:
            skipped += 1
            continue
        if len(row) >= 4:
            country_val = row[0].strip()
            quarter_val = row[1].strip().upper()
            name_val    = row[2].strip()
            try:
                budget_val = float(str(row[3]).replace(',','').replace('$','').strip() or 0)
            except:
                budget_val = 0
        else:
            country_val = request.form.get('country', '')
            quarter_val = row[0].strip().upper()
            name_val    = row[1].strip()
            try:
                budget_val = float(str(row[2]).replace(',','').replace('$','').strip() or 0)
            except:
                budget_val = 0

        if not country_val or not quarter_val or not name_val:
            skipped += 1
            continue

        if not quarter_val.startswith('Q'):
            quarter_val = 'Q' + quarter_val

        # Create budget record if it doesn't exist
        has_bud = any(r['country']==country_val and r['quarter']==quarter_val for r in existing_bud)
        if not has_bud:
            ws_bud.append_row([str(uuid.uuid4())[:8], country_val, quarter_val, 0, now])
            existing_bud.append({'country':country_val,'quarter':quarter_val,'total_budget':0})

        # Skip duplicate channel
        dup = any(r['country']==country_val and r['quarter']==quarter_val and r['name']==name_val for r in existing_ch)
        if dup:
            skipped += 1
            continue

        sort_order = len([r for r in existing_ch if r['country']==country_val and r['quarter']==quarter_val])
        ch_id = "ch_" + str(uuid.uuid4())[:8]
        ws_ch.append_row([ch_id, country_val, quarter_val, name_val, budget_val, sort_order, now])
        existing_ch.append({'id':ch_id,'country':country_val,'quarter':quarter_val,'name':name_val,'budget':budget_val,'sort_order':sort_order})
        saved += 1
        saved_rows.append({"country":country_val, "quarter":quarter_val, "name":name_val, "budget":budget_val})

    invalidate_cache(TAB_CHANNELS)
    invalidate_cache(TAB_BUDGETS)
    return jsonify({"ok": True, "saved": saved, "skipped": skipped, "rows": saved_rows})

# ── BULK BUDGET IMPORT ─────────────────────────────────────────
@app.route("/api/import/budgets", methods=["POST"])
@require_login
@require_admin
def api_import_budgets():
    """
    Accepts CSV or XLSX with columns: Country, Quarter, Total Budget
    Sets country-level total budgets in bulk.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files['file']
    fname = f.filename.lower()

    parsed_rows = []
    try:
        if fname.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)
            ws_xl = wb.active
            all_rows = list(ws_xl.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return jsonify({"error": "Empty file"}), 400
            first = [str(v or '').lower() for v in all_rows[0]]
            start = 1 if any(k in ' '.join(first) for k in ['country','quarter','budget']) else 0
            for row in all_rows[start:]:
                vals = [str(v).strip() if v is not None else '' for v in row]
                if any(vals):
                    parsed_rows.append(vals)
        else:
            content = f.read().decode('utf-8-sig')
            lines = [l.strip() for l in content.splitlines() if l.strip()]
            if not lines:
                return jsonify({"error": "Empty file"}), 400
            first = lines[0].lower()
            has_header = 'country' in first or 'quarter' in first or 'budget' in first
            data_lines = lines[1:] if has_header else lines
            import csv as _csv
            parsed_rows = list(_csv.reader(data_lines))
    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 400

    ws = get_sheet(TAB_BUDGETS)
    existing = ws.get_all_records()
    now = datetime.utcnow().isoformat()
    saved, skipped = 0, 0
    saved_rows = []

    for row in parsed_rows:
        if len(row) < 3:
            skipped += 1
            continue
        country_val = row[0].strip()
        quarter_val = row[1].strip().upper()
        if not quarter_val.startswith('Q'):
            quarter_val = 'Q' + quarter_val
        try:
            total = float(str(row[2]).replace(',','').replace('$','').strip() or 0)
        except:
            skipped += 1
            continue
        if not country_val or not quarter_val:
            skipped += 1
            continue

        idx = next((i for i,r in enumerate(existing) if r['country']==country_val and r['quarter']==quarter_val), None)
        if idx is not None:
            ws.update(f"A{idx+2}:E{idx+2}", [[existing[idx]['id'], country_val, quarter_val, total, now]])
            existing[idx]['total_budget'] = total
        else:
            ws.append_row([str(uuid.uuid4())[:8], country_val, quarter_val, total, now])
            existing.append({'country':country_val,'quarter':quarter_val,'total_budget':total})
        saved += 1
        saved_rows.append({"country":country_val, "quarter":quarter_val, "total":total})

    invalidate_cache(TAB_BUDGETS)
    return jsonify({"ok": True, "saved": saved, "skipped": skipped, "rows": saved_rows})

# ── CSV EXPORT ────────────────────────────────────────────────
@app.route("/api/export")
@require_login
def api_export():
    user = session["user"]
    rows = get_sheet(TAB_ENTRIES).get_all_records()
    if user != ADMIN_MARKET:
        rows = [r for r in rows if r["country"]==user]
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Country","Quarter","Month","Channel","Activity",
                "BU","Finance Category","Marketing Category","Description",
                "Planned (USD)","Confirmed (USD)","Actual (USD)",
                "JIRA","Vendor","Notes","Approved","Invoices","Entered By","Updated At"])
    for r in rows:
        inv_count = len(json.loads(r.get("invoice_names") or "[]"))
        w.writerow([
            r["country"], r["quarter"], r["month"], r["channel_name"],
            r.get("activity_name",""),
            r["bu"], r["finance_cat"], r["marketing_cat"], r["description"],
            r["planned"], r["confirmed"], r["actual"],
            r["jira"], r["vendor"], r["notes"],
            "Yes" if str(r["approved"])=="True" else "No",
            f"{inv_count} file(s)", r["entered_by"], r["updated_at"]
        ])
    fname = f"APAC_Budget_{'ALL' if user==ADMIN_MARKET else user}_{datetime.now().strftime('%Y-%m-%d')}.csv"
    return send_file(io.BytesIO(out.getvalue().encode("utf-8-sig")),
                     mimetype="text/csv", as_attachment=True, download_name=fname)

# ── ADMIN OVERVIEW ────────────────────────────────────────────
@app.route("/api/admin/overview")
@require_login
@require_admin
def api_admin_overview():
    budgets = get_sheet(TAB_BUDGETS).get_all_records()
    entries = get_sheet(TAB_ENTRIES).get_all_records()
    result = []
    for b in budgets:
        ces = [e for e in entries if e["country"]==b["country"] and e["quarter"]==b["quarter"]]
        result.append({
            "country": b["country"], "quarter": b["quarter"],
            "total": float(b["total_budget"] or 0),
            "planned": sum(float(e.get("planned") or 0) for e in ces),
            "actual":  sum(float(e.get("actual")  or 0) for e in ces),
            "entries": len(ces),
        })
    return jsonify(result)

# ── XLSX EXPORT (APAC budget format) ─────────────────────────
@app.route("/api/export/xlsx")
@require_login
def api_export_xlsx():
    from export_xlsx import build
    import tempfile, os

    user = session["user"]

    # Pull all entries + budget configs from Sheets
    all_entries  = get_sheet(TAB_ENTRIES).get_all_records()
    all_channels = get_sheet(TAB_CHANNELS).get_all_records()
    all_budgets  = get_sheet(TAB_BUDGETS).get_all_records()

    if user != ADMIN_MARKET:
        all_entries  = [e for e in all_entries  if e["country"] == user]
        all_channels = [c for c in all_channels if c["country"] == user]
        all_budgets  = [b for b in all_budgets  if b["country"] == user]

    # Build budget_data rows in the format expected by build()
    budget_rows = []
    for entry in all_entries:
        # Map channel name to category
        ch = next((c for c in all_channels if c["id"] == entry.get("channel_id", "")), {})
        monthly = [0] * 12
        # If month is set, place actual/planned in correct month slot
        month_val = entry.get("month", "")
        month_map = {
            "2025-07":0,"2025-08":1,"2025-09":2,"2025-10":3,
            "2025-11":4,"2025-12":5,"2026-01":6,"2026-02":7,
            "2026-03":8,"2026-04":9,"2026-05":10,"2026-06":11,
        }
        slot = month_map.get(month_val, -1)
        if slot >= 0:
            monthly[slot] = float(entry.get("planned") or 0)

        budget_rows.append({
            "account":       entry.get("bu", ""),
            "budget_owner":  entry.get("entered_by", ""),
            "category":      ch.get("name", entry.get("channel_name", "")),
            "country":       entry.get("country", ""),
            "vendor":        entry.get("vendor", ""),
            "note":          entry.get("description", ""),
            "fy26_budget":   float(entry.get("planned") or 0),
            "monthly":       monthly,
        })

    # Generate temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    build(tmp.name, budget_rows)

    fname = f"APAC_Budget_FY26_{'ALL' if user==ADMIN_MARKET else user}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response = send_file(tmp.name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)

    @response.call_on_close
    def cleanup():
        try: os.unlink(tmp.name)
        except: pass

    return response

if __name__ == "__main__":
    app.run(debug=True, port=5000)